"""
file  -  Workbook
pages - Worksheets

worksheets -> rows and cols
rows    -> numbered - 1 ... 1048576
cols    -> labeled A..Z, AA..AZ, BA..BZ.... XFD (16384)
cell    -> intersection of row and col
cell address => colname + rowno
                    A   +  10    = A10

worksheet - sheet1, sheet2, sheet3.....
we can rename a worksheet

"""


from openpyxl import Workbook
from datetime import datetime
wb = Workbook()

ws = wb.active

ws.title = "myworksheet"

ws['B5'] = 45000
ws['C5'] = "hello world"
ws['D5'] = datetime.now()

wb.save("FirstExcel.xlsx")