
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet('Employee')

wb.active = wb['Employee']

ws = wb.active

cell = ws['A5']

data = [
    ['empid', 'empname', 'age', 'dept', 'salary'],
    ['501', 'Jack', '28', 'HR', '65000'],
    ['235', 'Jill', '34', 'FIN', '35000'],
    ['150', 'John', '26', 'AC', '25000'],
    ['325', 'Jai', '21', 'DEV', '55000'],
    ['450', 'Guru', '38', 'DEV', '85000'],
    ['182', 'Sita', '32', 'PRC', '75000'],
    ['297', 'Gita', '36', 'SUP', '60000']
]

for row in data:
    ws.append(row)

wb.save("FirstExcel.xlsx")
