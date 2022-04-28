from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Product', 'Sales'),
    ('Pepsi', 350),
    ('Coke', 290),
    ('Sprite',230),
    ('Thumbs up', 425),
    ('Mirinda', 250),
    ('Maaza', 180),
    ('Slice', 220),
    ('7 up', 310)
]

print(f"data :{data}")

for row in data:
    ws.append(row)

print(ws.dimensions)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=9)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=9)

chart = BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.x_axis.title = "Products"
chart.y_axis.title = 'Sales in lakhs'
chart.title = "Baverage Sales"

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")