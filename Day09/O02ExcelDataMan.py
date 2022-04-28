
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Employee']

ws = wb.active

print(ws.dimensions)
# A6 : E13          =>  col1... col5, row6...r13
for row in ws.iter_rows(min_row=6, min_col=2, max_row=13, max_col=2):
    for cell in row:
        if cell.value == "John":
            cell.value = "john travolta".upper()

wb.save("FirstExcel.xlsx")
