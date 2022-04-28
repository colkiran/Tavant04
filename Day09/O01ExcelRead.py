
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

print(f"Worksheet Names :{wb.sheetnames}")

wb.active = wb['Employee']

ws = wb.active

print(ws.dimensions)

dataRange = ws[ws.dimensions]

for c1, c2, c3, c4, c5 in dataRange:
    print("{0:5}\t{1:10}\t{2:5}\t{3:5}\t{4:5}".format( c1.value, c2.value, c3.value, c4.value, c5.value ))

